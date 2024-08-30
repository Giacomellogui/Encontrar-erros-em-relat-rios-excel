import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import json
from spellchecker import SpellChecker

# Inicializa o dicionário de títulos esperados
expected_titles = {}

# Inicializa o corretor ortográfico
spell = SpellChecker(language='pt')

# Função para verificar os títulos
def analisar_titulos(ws, expected_title):
    title_found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.font.bold:
                if cell.value.strip().lower() == expected_title.lower():
                    title_found = True
                    break
        if title_found:
            break
    return title_found

# Função para verificar as perguntas
def analisar_perguntas(ws, expected_question):
    question_found = False
    expected_question = expected_question.strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_value = cell.value.strip().lower()
                if cell_value.startswith("pergunta") and expected_question in cell_value:
                    question_found = True
                    break
        if question_found:
            break
    return question_found

# Função para verificar a grafia em um texto
def verificar_grafia(text):
    palavras_erradas = spell.unknown(text.split())
    return palavras_erradas

# Função para adicionar títulos e perguntas à lista
def adicionar_titulo():
    tabela = selected_sheet.get().strip()
    titulos = titulos_entry.get().strip()
    pergunta = pergunta_entry.get().strip()

    if tabela:
        if tabela not in expected_titles:
            expected_titles[tabela] = {"titulos": [], "pergunta": ""}
        if titulos:
            expected_titles[tabela]["titulos"].append(titulos)
        if pergunta:
            expected_titles[tabela]["pergunta"] = pergunta
        lista_titulos.insert(tk.END, f"{tabela}: {titulos} | Pergunta: {pergunta}")
    else:
        messagebox.showwarning("Aviso", "Por favor, insira ao menos um título ou uma pergunta.")

    titulos_entry.delete(0, tk.END)
    pergunta_entry.delete(0, tk.END)

# Função para limpar os títulos e perguntas esperadas
def limpar_titulos():
    expected_titles.clear()
    lista_titulos.delete(0, tk.END)
    messagebox.showinfo("Informação", "Todos os títulos e perguntas esperados foram removidos.")

# Função para realizar a análise de títulos
def realizar_analise_titulos():
    if not file_path.get():
        messagebox.showwarning("Aviso", "Por favor, carregue um arquivo Excel.")
        return

    arquivo_excel = file_path.get()
    wb = load_workbook(arquivo_excel)
    abas_disponiveis = wb.sheetnames
    erros = {}

    for tabela, conteudo in expected_titles.items():
        if tabela not in abas_disponiveis:
            erros[tabela] = {"faltando": "Tabela não encontrada", "titulo": False}
            continue

        ws = wb[tabela]
        for titulo_esperado in conteudo["titulos"]:
            titulo_encontrado = analisar_titulos(ws, titulo_esperado)
            if not titulo_encontrado:
                erros.setdefault(tabela, {})["titulo"] = False
            else:
                # Verificar grafia dos títulos encontrados
                palavras_erradas = verificar_grafia(titulo_esperado)
                if palavras_erradas:
                    erros.setdefault(tabela, {})["grafia_titulo"] = palavras_erradas

    # Exibir os erros de títulos
    resultado_texto.delete(1.0, tk.END)
    if erros:
        result_text = "Relatório de Erros (Títulos):\n\n"
        for tabela, problema in erros.items():
            result_text += f"Tabela '{tabela}':\n"
            if "titulo" in problema and not problema["titulo"]:
                result_text += " - Título esperado não encontrado ou não está em negrito.\n"
            if "grafia_titulo" in problema:
                result_text += f" - Erros de grafia no título: {', '.join(problema['grafia_titulo'])}\n"
            result_text += "\n"
        resultado_texto.insert(tk.END, result_text)
    else:
        resultado_texto.insert(tk.END, "Nenhum erro encontrado nos títulos!")

# Função para realizar a análise de perguntas
def realizar_analise_perguntas():
    if not file_path.get():
        messagebox.showwarning("Aviso", "Por favor, carregue um arquivo Excel.")
        return

    arquivo_excel = file_path.get()
    wb = load_workbook(arquivo_excel)
    abas_disponiveis = wb.sheetnames
    erros = {}

    for tabela, conteudo in expected_titles.items():
        if tabela not in abas_disponiveis:
            erros[tabela] = {"faltando": "Tabela não encontrada", "pergunta": False}
            continue

        ws = wb[tabela]
        pergunta_esperada = conteudo["pergunta"]
        if pergunta_esperada:
            pergunta_encontrada = analisar_perguntas(ws, pergunta_esperada)
            if not pergunta_encontrada:
                erros.setdefault(tabela, {})["pergunta"] = False
            else:
                # Verificar grafia das perguntas encontradas
                palavras_erradas = verificar_grafia(pergunta_esperada)
                if palavras_erradas:
                    erros.setdefault(tabela, {})["grafia_pergunta"] = palavras_erradas

    # Exibir os erros de perguntas
    resultado_texto.delete(1.0, tk.END)
    if erros:
        result_text = "Relatório de Erros (Perguntas):\n\n"
        for tabela, problema in erros.items():
            result_text += f"Tabela '{tabela}':\n"
            if "pergunta" in problema and not problema["pergunta"]:
                result_text += " - Pergunta esperada não encontrada.\n"
            if "grafia_pergunta" in problema:
                result_text += f" - Erros de grafia na pergunta: {', '.join(problema['grafia_pergunta'])}\n"
            result_text += "\n"
        resultado_texto.insert(tk.END, result_text)
    else:
        resultado_texto.insert(tk.END, "Nenhum erro encontrado nas perguntas!")

# Função para carregar o arquivo Excel e preencher o menu de opções de abas
def carregar_arquivo():
    filename = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
    )
    if filename:
        file_path.set(filename)
        wb = load_workbook(filename)
        abas_disponiveis = wb.sheetnames
        selected_sheet.set(abas_disponiveis[0])  # Define a primeira aba como padrão
        sheet_menu['menu'].delete(0, 'end')
        for aba in abas_disponiveis:
            sheet_menu['menu'].add_command(label=aba, command=tk._setit(selected_sheet, aba))
        messagebox.showinfo("Abas Disponíveis", f"Abas encontradas: {', '.join(abas_disponiveis)}")

# Função para salvar as configurações em um arquivo JSON
def salvar_configuracoes():
    if not expected_titles:
        messagebox.showwarning("Aviso", "Não há títulos e perguntas para salvar.")
        return
    
    save_path = filedialog.asksaveasfilename(
        defaultextension=".json",
        filetypes=[("Arquivos JSON", "*.json"), ("Todos os arquivos", "*.*")],
        title="Salvar Configurações"
    )
    
    if save_path:
        try:
            with open(save_path, 'w', encoding='utf-8') as file:
                json.dump(expected_titles, file, ensure_ascii=False, indent=4)
            messagebox.showinfo("Sucesso", "Configurações salvas com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar configurações: {e}")

# Função para carregar as configurações de um arquivo JSON
def carregar_configuracoes():
    load_path = filedialog.askopenfilename(
        filetypes=[("Arquivos JSON", "*.json"), ("Todos os arquivos", "*.*")],
        title="Carregar Configurações"
    )
    
    if load_path:
        try:
            with open(load_path, 'r', encoding='utf-8') as file:
                global expected_titles
                expected_titles = json.load(file)
                
            lista_titulos.delete(0, tk.END)
            for tabela, conteudo in expected_titles.items():
                for titulo in conteudo["titulos"]:
                    lista_titulos.insert(tk.END, f"{tabela}: {titulo} | Pergunta: {conteudo['pergunta']}")
                    
            messagebox.showinfo("Sucesso", "Configurações carregadas com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar configurações: {e}")

# Inicialização da interface gráfica
root = tk.Tk()
root.title("Verificação de Títulos e Perguntas no Excel")

file_path = tk.StringVar()
selected_sheet = tk.StringVar()

frame_arquivo = tk.Frame(root)
frame_arquivo.pack(padx=10, pady=10)

lbl_caminho = tk.Label(frame_arquivo, text="Caminho do Arquivo:")
lbl_caminho.pack(side=tk.LEFT)

entry_caminho = tk.Entry(frame_arquivo, textvariable=file_path, width=50)
entry_caminho.pack(side=tk.LEFT, padx=(0, 10))

btn_carregar = tk.Button(frame_arquivo, text="Carregar", command=carregar_arquivo)
btn_carregar.pack(side=tk.LEFT)

frame_opcoes = tk.Frame(root)
frame_opcoes.pack(padx=10, pady=10)

lbl_sheet = tk.Label(frame_opcoes, text="Selecione a Tabela:")
lbl_sheet.pack(side=tk.LEFT)

sheet_menu = ttk.OptionMenu(frame_opcoes, selected_sheet, "")
sheet_menu.pack(side=tk.LEFT)

frame_titulos = tk.Frame(root)
frame_titulos.pack(padx=10, pady=10)

lbl_titulos = tk.Label(frame_titulos, text="Títulos:")
lbl_titulos.grid(row=0, column=0)

titulos_entry = tk.Entry(frame_titulos, width=50)
titulos_entry.grid(row=0, column=1)

lbl_pergunta = tk.Label(frame_titulos, text="Pergunta:")
lbl_pergunta.grid(row=1, column=0)

pergunta_entry = tk.Entry(frame_titulos, width=50)
pergunta_entry.grid(row=1, column=1)

btn_adicionar = tk.Button(frame_titulos, text="Adicionar Título e Pergunta", command=adicionar_titulo)
btn_adicionar.grid(row=2, columnspan=2, pady=5)

frame_lista = tk.Frame(root)
frame_lista.pack(padx=10, pady=10)

lista_titulos = tk.Listbox(frame_lista, width=80)
lista_titulos.pack()

frame_botoes = tk.Frame(root)
frame_botoes.pack(padx=10, pady=10)

btn_analisar_titulos = tk.Button(frame_botoes, text="Analisar Títulos", command=realizar_analise_titulos)
btn_analisar_titulos.grid(row=0, column=0, padx=5)

btn_analisar_perguntas = tk.Button(frame_botoes, text="Analisar Perguntas", command=realizar_analise_perguntas)
btn_analisar_perguntas.grid(row=0, column=1, padx=5)

btn_salvar_config = tk.Button(frame_botoes, text="Salvar Configurações", command=salvar_configuracoes)
btn_salvar_config.grid(row=0, column=2, padx=5)

btn_carregar_config = tk.Button(frame_botoes, text="Carregar Configurações", command=carregar_configuracoes)
btn_carregar_config.grid(row=0, column=3, padx=5)

btn_limpar_titulos = tk.Button(frame_botoes, text="Limpar Títulos e Perguntas", command=limpar_titulos)
btn_limpar_titulos.grid(row=0, column=4, padx=5)

frame_resultado = tk.Frame(root)
frame_resultado.pack(padx=10, pady=10)

resultado_texto = tk.Text(frame_resultado, height=15, width=80)
resultado_texto.pack()

root.mainloop()
